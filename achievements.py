"""
Achievements data extractor

src:
achievements.xlsx https://docs.qq.com/sheet/DS01hbnZwZm5KVnBB?tab=BB08J2
achievements.py https://github.com/KimigaiiWuyi/GenshinUID/blob/main/GenshinUID/tools/get_achievement_json.py
"""

import json
import warnings
from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
import_path = "src/achievements.xlsx"
export_path = Path("Resources")

wb: Workbook = load_workbook(import_path)
ws_daily: Worksheet = wb['成就相关每日委托']
ws_all: Worksheet = wb['正式服成就汇总']

result_daily = []
result_all = []


class BaseAchievement(BaseModel):
    name: str
    desc: str
    guide: str
    link: str


class Achievement(BaseAchievement):
    book: str


class TaskAchievement(BaseAchievement):
    task: List[str]


def load_daily_achievements():
    is_first = False
    for row in range(3, 100):
        ach = TaskAchievement(
            task=[],
            name=ws_daily.cell(row, 4).value or "",
            desc=ws_daily.cell(row, 5).value or "",
            guide=ws_daily.cell(row, 6).value or "",
            link=ws_daily.cell(row, 6).hyperlink.target if ws_daily.cell(row, 6).hyperlink else '',
        )
        task = ws_daily.cell(row, 3).value or ""
        if not task:
            if is_first:
                break
            is_first = True
            continue
        else:
            is_first = False
        ach.task = [i for i in task.split('\n') if not i.startswith('(')]
        result_daily.append(ach.dict())


def load_all_achievements(book: Worksheet, loop: int, bn: int, an: int, dn: int, gn: int):
    for row in range(3, loop):
        ach = Achievement(
            book=book.cell(row, bn).value or "",
            name=book.cell(row, an).value or "",
            desc=book.cell(row, dn).value or "",
            guide=book.cell(row, gn).value or "",
            link=book.cell(row, gn).hyperlink.target if book.cell(row, gn).hyperlink else '',
        )
        if not ach.book:
            break
        result_all.append(ach.dict())


def save_achievements():
    export_path.mkdir(parents=True, exist_ok=True)
    with open(export_path / 'achievements_daily.json', 'w', encoding='utf-8') as f:
        json.dump(result_daily, f, indent=4, ensure_ascii=False)
    with open(export_path / 'achievements_all.json', 'w', encoding='utf-8') as f:
        json.dump(result_all, f, indent=4, ensure_ascii=False)


if __name__ == '__main__':
    load_all_achievements(ws_all, 1000, 5, 6, 7, 11)
    load_daily_achievements()
    save_achievements()
